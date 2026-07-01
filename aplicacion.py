import tkinter as tk
from tkinter import messagebox, ttk, filedialog
from datetime import datetime
from tkcalendar import DateEntry
from openpyxl import Workbook

from Dao.DAOcliente import DAOCliente
from Dao.DAOdestino import DAODestino
from Dao.DAOpaquete import DAOPaquete
from Dao.DAOreserva import DAOReserva
from clases.cliente import Cliente
from clases.destino import Destino
from clases.Paquete_Turistico import PaqueteTuristico
from clases.reserva import Reserva
from seguridad import Seguridad
from conexiones import Conexion

# Paleta Verde
FONDO_PRINCIPAL = "#f0fdf4"
MENU_BG         = "#14532d"
ACCENT          = "#4ade80"
ACCENT_DARK     = "#22c55e"
TEXTO           = "#052e16"
TEXTO_CLARO     = "#86efac"
GRIS            = "#4b5563"
ROJO            = "#ef4444"

class App:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Viajes Aventura - Sistema de Gestión")
        self.root.geometry("1450x820")
        self.root.configure(bg=FONDO_PRINCIPAL)

        self.usuario_actual = None
        self.rol_actual = None

        self.sidebar = tk.Frame(self.root, bg=MENU_BG, width=280)
        self.sidebar.place(x=0, y=0, relheight=1)
        self.sidebar.pack_propagate(False)

        self.contenido = tk.Frame(self.root, bg=FONDO_PRINCIPAL)
        self.contenido.place(x=280, y=0, relheight=1, relwidth=1, width=-280)

        self.mostrar_login()

    def _limpiar_contenido(self):
        for w in self.contenido.winfo_children():
            w.destroy()

    def _limpiar_sidebar(self):
        for w in self.sidebar.winfo_children():
            w.destroy()

    def _boton_menu(self, texto: str, comando, icono="🌿"):
        btn = tk.Button(self.sidebar, text=f"{icono}  {texto}", command=comando,
                        bg=MENU_BG, fg=TEXTO_CLARO, activebackground=ACCENT,
                        relief="flat", font=("Arial", 12, "bold"), anchor="w",
                        padx=30, height=2, cursor="hand2", bd=0)
        btn.pack(fill="x", pady=4)

    # ==================== LOGIN ====================
    def mostrar_login(self):
        self._limpiar_sidebar()
        self._limpiar_contenido()
        frame = tk.Frame(self.contenido, bg=FONDO_PRINCIPAL, padx=40, pady=30)
        frame.place(relx=0.5, rely=0.48, anchor="center")

        tk.Label(frame, text="🌍 Viajes Aventura", bg=FONDO_PRINCIPAL, fg=ACCENT_DARK, font=("Arial", 36, "bold")).pack(pady=(0,8))
        tk.Label(frame, text="Tu próxima aventura comienza aquí", bg=FONDO_PRINCIPAL, fg=GRIS, font=("Arial", 13)).pack(pady=(0,24))

        form_frame = tk.Frame(frame, bg=FONDO_PRINCIPAL)
        form_frame.pack(fill="x")

        tk.Label(form_frame, text="Usuario / RUT", bg=FONDO_PRINCIPAL, fg=TEXTO, font=("Arial", 11, "bold")).grid(row=0, column=0, sticky="w")
        entry_user = tk.Entry(form_frame, width=40, font=("Arial", 11), relief="solid", bd=2)
        entry_user.grid(row=1, column=0, pady=8, sticky="ew")

        tk.Label(form_frame, text="Contraseña", bg=FONDO_PRINCIPAL, fg=TEXTO, font=("Arial", 11, "bold")).grid(row=2, column=0, sticky="w")
        password_frame = tk.Frame(form_frame, bg=FONDO_PRINCIPAL)
        password_frame.grid(row=3, column=0, pady=8, sticky="ew")
        entry_pw = tk.Entry(password_frame, width=33, font=("Arial", 11), show="*", relief="solid", bd=2)
        entry_pw.pack(side="left", fill="x", expand=True)

        show_password = {"active": False}
        def toggle_password():
            show_password["active"] = not show_password["active"]
            entry_pw.config(show="" if show_password["active"] else "*")
            btn_show.config(text="Ocultar" if show_password["active"] else "Mostrar")

        btn_show = tk.Button(password_frame, text="Mostrar", bg="#e2e8f0", fg=TEXTO, relief="solid", width=10, command=toggle_password)
        btn_show.pack(side="left", padx=(8,0))

        self.lbl_estado = tk.Label(frame, text="", bg=FONDO_PRINCIPAL, fg=ROJO, font=("Arial", 10))
        self.lbl_estado.pack(pady=(4,16), anchor="w")

        buttons_frame = tk.Frame(frame, bg=FONDO_PRINCIPAL)
        buttons_frame.pack(fill="x")
        tk.Button(buttons_frame, text="INICIAR SESIÓN", bg=ACCENT_DARK, fg="white", font=("Arial", 13, "bold"), width=38, height=2,
                  command=lambda: self._procesar_login(entry_user.get().strip(), entry_pw.get())).pack(pady=4)
        tk.Button(buttons_frame, text="REGISTRARSE", bg="#e2e8f0", fg=TEXTO, font=("Arial", 12, "bold"), width=38, height=1,
                  command=self._mostrar_registro).pack(pady=4)

    def _procesar_login(self, usuario, clave):
        self.lbl_estado.config(text="")

        if usuario == "12345" and clave == "ADMIN123":
            from clases.cliente import Cliente
            admin = Cliente(1, "Administrador", "12345", "", "", "")
            self._login_exitoso(admin, "Administrador")
            return

        try:
            cliente = DAOCliente().buscar(usuario)
            if cliente and Seguridad.validar_clave(clave, cliente.get_contrasena()):
                self._login_exitoso(cliente, "Cliente")
                return
        except:
            pass

        self.lbl_estado.config(text="✘ Credenciales incorrectas", fg=ROJO)

    def _login_exitoso(self, usuario, rol):
        self.usuario_actual = usuario
        self.rol_actual = rol
        if rol == "Administrador":
            self._panel_admin()
        else:
            self._panel_cliente()

    def _mostrar_registro(self):
        self._limpiar_sidebar()
        self._limpiar_contenido()
        frame = tk.Frame(self.contenido, bg=FONDO_PRINCIPAL)
        frame.place(relx=0.5, rely=0.45, anchor="center")

        tk.Label(frame, text="Registrar Cliente", bg=FONDO_PRINCIPAL, fg=ACCENT_DARK, font=("Arial", 32, "bold")).pack(pady=10)
        tk.Label(frame, text="Crea tu cuenta para ver paquetes y reservas como cliente.", bg=FONDO_PRINCIPAL, fg=GRIS, font=("Arial", 12)).pack(pady=8)

        tk.Label(frame, text="Nombre", bg=FONDO_PRINCIPAL, fg=TEXTO, font=("Arial", 11, "bold")).pack(anchor="w", padx=40)
        nombre_entry = tk.Entry(frame, width=36, font=("Arial", 11), relief="solid", bd=2)
        nombre_entry.pack(pady=6, ipady=8, padx=40)

        tk.Label(frame, text="RUT", bg=FONDO_PRINCIPAL, fg=TEXTO, font=("Arial", 11, "bold")).pack(anchor="w", padx=40)
        rut_entry = tk.Entry(frame, width=36, font=("Arial", 11), relief="solid", bd=2)
        rut_entry.pack(pady=6, ipady=8, padx=40)

        tk.Label(frame, text="Correo", bg=FONDO_PRINCIPAL, fg=TEXTO, font=("Arial", 11, "bold")).pack(anchor="w", padx=40)
        correo_entry = tk.Entry(frame, width=36, font=("Arial", 11), relief="solid", bd=2)
        correo_entry.pack(pady=6, ipady=8, padx=40)

        tk.Label(frame, text="Teléfono", bg=FONDO_PRINCIPAL, fg=TEXTO, font=("Arial", 11, "bold")).pack(anchor="w", padx=40)
        telefono_entry = tk.Entry(frame, width=36, font=("Arial", 11), relief="solid", bd=2)
        telefono_entry.pack(pady=6, ipady=8, padx=40)

        tk.Label(frame, text="Contraseña", bg=FONDO_PRINCIPAL, fg=TEXTO, font=("Arial", 11, "bold")).pack(anchor="w", padx=40)
        contrasena_entry = tk.Entry(frame, width=36, font=("Arial", 11), show="*", relief="solid", bd=2)
        contrasena_entry.pack(pady=6, ipady=8, padx=40)

        tk.Label(frame, text="Confirmar Contraseña", bg=FONDO_PRINCIPAL, fg=TEXTO, font=("Arial", 11, "bold")).pack(anchor="w", padx=40)
        confirmar_entry = tk.Entry(frame, width=36, font=("Arial", 11), show="*", relief="solid", bd=2)
        confirmar_entry.pack(pady=6, ipady=8, padx=40)

        def registrar_cliente():
            nombre = nombre_entry.get().strip()
            rut = rut_entry.get().strip()
            correo = correo_entry.get().strip()
            telefono = telefono_entry.get().strip()
            contrasena = contrasena_entry.get().strip()
            confirmar = confirmar_entry.get().strip()

            if not nombre or not rut or not correo or not telefono or not contrasena or not confirmar:
                messagebox.showwarning("Registrar", "Todos los campos son obligatorios.")
                return
            if contrasena != confirmar:
                messagebox.showwarning("Registrar", "Las contraseñas no coinciden.")
                return
            if DAOCliente().buscar(rut):
                messagebox.showwarning("Registrar", "Ya existe un cliente con ese RUT.")
                return

            hash_clave = Seguridad.encriptar_clave(contrasena)
            cliente = Cliente(None, nombre, rut, correo, hash_clave, telefono)
            if DAOCliente().registrar(cliente):
                messagebox.showinfo("Registro", "Cuenta creada. Ahora puedes iniciar sesión.")
                self.mostrar_login()
            else:
                messagebox.showerror("Registro", "No se pudo crear la cuenta.")

        tk.Button(frame, text="REGISTRARSE", bg=ACCENT_DARK, fg="white", font=("Arial", 13, "bold"), width=30, height=2,
                  command=registrar_cliente).pack(pady=10)
        tk.Button(frame, text="Volver al Login", bg="#e2e8f0", fg=TEXTO, font=("Arial", 11, "bold"), width=30, height=1,
                  command=self.mostrar_login).pack()

    def _panel_admin(self):
        self._limpiar_sidebar()
        tk.Label(self.sidebar, text="🌍 Viajes Aventura", bg=MENU_BG, fg=TEXTO_CLARO, font=("Arial", 22, "bold")).pack(pady=(30,5))
        tk.Label(self.sidebar, text="Administrador", bg=MENU_BG, fg=ACCENT, font=("Arial", 10, "bold")).pack(pady=(0,20))

        self._boton_menu("Destinos", self._vista_destinos, "🏝️")
        self._boton_menu("Paquetes", self._vista_paquetes, "🎒")
        self._boton_menu("Reservas", self._vista_reservas, "📋")
        self._boton_menu("Exportar", self._vista_exportar, "📊")
        self._boton_menu("Cerrar Sesión", self._cerrar_sesion, "🚪")

        self._vista_destinos()

    def _cerrar_sesion(self):
        self.usuario_actual = None
        self.rol_actual = None
        self.mostrar_login()

    # ==================== VISTAS FUNCIONALES ====================
    def _vista_destinos(self):
        self._limpiar_contenido()
        tk.Label(self.contenido, text="Gestión de Destinos", bg=FONDO_PRINCIPAL, fg=TEXTO, font=("Arial", 28, "bold")).pack(anchor="w", padx=50, pady=30)

        btn_frame = tk.Frame(self.contenido, bg=FONDO_PRINCIPAL)
        btn_frame.pack(fill="x", padx=50, pady=10)
        tk.Button(btn_frame, text="➕ Nuevo Destino", bg=ACCENT_DARK, fg="white", command=self._nuevo_destino, width=15).pack(side="left", padx=5)
        tk.Button(btn_frame, text="✏️ Editar Destino", bg=ACCENT, fg="white", command=self._editar_destino, width=15).pack(side="left", padx=5)
        tk.Button(btn_frame, text="🗑️ Eliminar Destino", bg=ROJO, fg="white", command=self._eliminar_destino, width=15).pack(side="left", padx=5)

        columns = ("ID", "Nombre", "Descripción", "Costo")
        tree_frame = tk.Frame(self.contenido, bg=FONDO_PRINCIPAL)
        tree_frame.pack(fill="both", expand=True, padx=50, pady=10)

        self.destino_tree = ttk.Treeview(tree_frame, columns=columns, show="headings", selectmode="browse")
        for col in columns:
            self.destino_tree.heading(col, text=col)
            self.destino_tree.column(col, width=200, anchor="center")
        self.destino_tree.pack(side="left", fill="both", expand=True)

        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.destino_tree.yview)
        self.destino_tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")

        self._cargar_destinos()

    def _cargar_destinos(self):
        self.destino_tree.delete(*self.destino_tree.get_children())
        for destino in DAODestino().obtener_todo():
            self.destino_tree.insert("", "end", values=(destino.get_id(), destino.get_nombre(), destino.get_descripcion(), destino.get_costo()))

    def _nuevo_destino(self):
        self._abrir_form_destino()

    def _editar_destino(self):
        seleccionado = self.destino_tree.selection()
        if not seleccionado:
            messagebox.showwarning("Editar Destino", "Selecciona un destino para editar.")
            return
        item = self.destino_tree.item(seleccionado)
        valores = item["values"]
        destino = Destino(valores[0], valores[1], valores[2], "", valores[3])
        self._abrir_form_destino(destino)

    def _eliminar_destino(self):
        seleccionado = self.destino_tree.selection()
        if not seleccionado:
            messagebox.showwarning("Eliminar Destino", "Selecciona un destino para eliminar.")
            return
        item = self.destino_tree.item(seleccionado)
        destino_id = item["values"][0]
        if messagebox.askyesno("Eliminar Destino", "¿Estás seguro de eliminar este destino?"):
            if DAODestino().eliminar(destino_id):
                messagebox.showinfo("Destino", "Destino eliminado correctamente.")
                self._cargar_destinos()
            else:
                messagebox.showerror("Destino", "No se pudo eliminar el destino.")

    def _abrir_form_destino(self, destino=None):
        ventana = tk.Toplevel(self.root)
        ventana.title("Destino")
        ventana.transient(self.root)
        ventana.grab_set()
        ventana.geometry("560x440")
        ventana.configure(bg=FONDO_PRINCIPAL)

        tk.Label(ventana, text="Destino", bg=FONDO_PRINCIPAL, fg=TEXTO, font=("Arial", 20, "bold")).pack(pady=12)

        frame = tk.Frame(ventana, bg=FONDO_PRINCIPAL)
        frame.pack(fill="both", expand=True, padx=20, pady=10)

        tk.Label(frame, text="Nombre", bg=FONDO_PRINCIPAL, fg=TEXTO).grid(row=0, column=0, sticky="w", pady=8)
        nombre_entry = tk.Entry(frame, width=50)
        nombre_entry.grid(row=0, column=1, pady=8, padx=10)

        tk.Label(frame, text="Descripción", bg=FONDO_PRINCIPAL, fg=TEXTO).grid(row=1, column=0, sticky="w", pady=8)
        descripcion_entry = tk.Entry(frame, width=50)
        descripcion_entry.grid(row=1, column=1, pady=8, padx=10)

        tk.Label(frame, text="Actividades", bg=FONDO_PRINCIPAL, fg=TEXTO).grid(row=2, column=0, sticky="w", pady=8)
        actividades_entry = tk.Entry(frame, width=50)
        actividades_entry.grid(row=2, column=1, pady=8, padx=10)

        tk.Label(frame, text="Costo", bg=FONDO_PRINCIPAL, fg=TEXTO).grid(row=3, column=0, sticky="w", pady=8)
        costo_entry = tk.Entry(frame, width=50)
        costo_entry.grid(row=3, column=1, pady=8, padx=10)

        if destino:
            nombre_entry.insert(0, destino.get_nombre())
            descripcion_entry.insert(0, destino.get_descripcion())
            actividades_entry.insert(0, destino.get_actividades())
            costo_entry.insert(0, destino.get_costo())

        def guardar():
            nombre = nombre_entry.get().strip()
            descripcion = descripcion_entry.get().strip()
            actividades = actividades_entry.get().strip()
            costo = costo_entry.get().strip()
            if not nombre or not costo:
                messagebox.showwarning("Destino", "Nombre y costo son obligatorios.")
                return
            try:
                costo_val = float(costo)
            except ValueError:
                messagebox.showwarning("Destino", "Costo debe ser un número válido.")
                return

            if destino:
                destino.set_nombre(nombre)
                destino.set_descripcion(descripcion)
                destino.set_actividades(actividades)
                destino.set_costo(costo_val)
                if DAODestino().actualizar(destino):
                    messagebox.showinfo("Destino", "Destino actualizado correctamente.")
                    self._cargar_destinos()
                    ventana.destroy()
                else:
                    messagebox.showerror("Destino", "No se pudo actualizar el destino.")
            else:
                nuevo_destino = Destino(None, nombre, descripcion, actividades, costo_val)
                if DAODestino().registrar(nuevo_destino):
                    messagebox.showinfo("Destino", "Destino creado correctamente.")
                    self._cargar_destinos()
                    ventana.destroy()
                else:
                    messagebox.showerror("Destino", "No se pudo crear el destino.")

        tk.Button(ventana, text="Guardar", bg=ACCENT_DARK, fg="white", command=guardar, width=14, height=2).pack(pady=18)

    def _vista_paquetes(self):
        self._limpiar_contenido()
        tk.Label(self.contenido, text="Ofertas de Viaje", bg=FONDO_PRINCIPAL, fg=TEXTO, font=("Arial", 28, "bold")).pack(anchor="w", padx=50, pady=30)

        btn_frame = tk.Frame(self.contenido, bg=FONDO_PRINCIPAL)
        btn_frame.pack(fill="x", padx=50, pady=10)
        tk.Button(btn_frame, text="➕ Nueva Oferta", bg=ACCENT_DARK, fg="white", command=self._nuevo_paquete, width=15).pack(side="left", padx=5)
        tk.Button(btn_frame, text="✏️ Editar Oferta", bg=ACCENT, fg="white", command=self._editar_paquete, width=15).pack(side="left", padx=5)
        tk.Button(btn_frame, text="🗑️ Eliminar Oferta", bg=ROJO, fg="white", command=self._eliminar_paquete, width=15).pack(side="left", padx=5)

        columns = ("ID", "Oferta", "Descripción", "Precio")
        tree_frame = tk.Frame(self.contenido, bg=FONDO_PRINCIPAL)
        tree_frame.pack(fill="both", expand=True, padx=50, pady=10)

        self.paquete_tree = ttk.Treeview(tree_frame, columns=columns, show="headings", selectmode="browse")
        for col in columns:
            self.paquete_tree.heading(col, text=col)
            self.paquete_tree.column(col, width=240 if col != "ID" else 80, anchor="center")
        self.paquete_tree.pack(side="left", fill="both", expand=True)

        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.paquete_tree.yview)
        self.paquete_tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")

        self._cargar_paquetes()

    def _cargar_paquetes(self):
        self.paquete_tree.delete(*self.paquete_tree.get_children())
        for paquete in DAOPaquete().obtener_todo():
            self.paquete_tree.insert("", "end", values=(paquete.get_id(), paquete.get_nombre(), paquete.get_descripcion(), paquete.get_precio()))

    def _nuevo_paquete(self):
        self._abrir_form_paquete()

    def _editar_paquete(self):
        seleccionado = self.paquete_tree.selection()
        if not seleccionado:
            messagebox.showwarning("Editar Paquete", "Selecciona un paquete para editar.")
            return
        item = self.paquete_tree.item(seleccionado)
        paquete_id = item["values"][0]
        paquete = DAOPaquete().obtener_por_id(paquete_id)
        if not paquete:
            messagebox.showerror("Editar Paquete", "No se pudo cargar el paquete seleccionado.")
            return
        self._abrir_form_paquete(paquete)

    def _eliminar_paquete(self):
        seleccionado = self.paquete_tree.selection()
        if not seleccionado:
            messagebox.showwarning("Eliminar Paquete", "Selecciona un paquete para eliminar.")
            return
        item = self.paquete_tree.item(seleccionado)
        paquete_id = item["values"][0]
        if messagebox.askyesno("Eliminar Paquete", "¿Estás seguro de eliminar este paquete?"):
            if DAOPaquete().eliminar(paquete_id):
                messagebox.showinfo("Paquete", "Paquete eliminado correctamente.")
                self._cargar_paquetes()
            else:
                messagebox.showerror("Paquete", "No se pudo eliminar el paquete.")

    def _abrir_form_paquete(self, paquete=None):
        ventana = tk.Toplevel(self.root)
        ventana.title("Paquete Turístico")
        ventana.transient(self.root)
        ventana.grab_set()
        ventana.geometry("620x620")
        ventana.configure(bg=FONDO_PRINCIPAL)

        tk.Label(ventana, text="Paquete Turístico", bg=FONDO_PRINCIPAL, fg=TEXTO, font=("Arial", 20, "bold")).pack(pady=12)

        frame = tk.Frame(ventana, bg=FONDO_PRINCIPAL)
        frame.pack(fill="both", expand=True, padx=20, pady=10)

        tk.Label(frame, text="Nombre de la oferta", bg=FONDO_PRINCIPAL, fg=TEXTO).grid(row=0, column=0, sticky="w", pady=8)
        nombre_entry = tk.Entry(frame, width=50)
        nombre_entry.grid(row=0, column=1, pady=8, padx=10)

        tk.Label(frame, text="Descripción de la oferta", bg=FONDO_PRINCIPAL, fg=TEXTO).grid(row=1, column=0, sticky="w", pady=8)
        descripcion_entry = tk.Entry(frame, width=50)
        descripcion_entry.grid(row=1, column=1, pady=8, padx=10)

        tk.Label(frame, text="Precio único", bg=FONDO_PRINCIPAL, fg=TEXTO).grid(row=2, column=0, sticky="w", pady=8)
        precio_entry = tk.Entry(frame, width=50)
        precio_entry.grid(row=2, column=1, pady=8, padx=10)

        tk.Label(frame, text="Fecha Inicio (opcional)", bg=FONDO_PRINCIPAL, fg=TEXTO).grid(row=3, column=0, sticky="w", pady=8)
        inicio_entry = DateEntry(frame, width=47, date_pattern='yyyy-mm-dd')
        inicio_entry.grid(row=3, column=1, pady=8, padx=10)

        tk.Label(frame, text="Fecha Fin (opcional)", bg=FONDO_PRINCIPAL, fg=TEXTO).grid(row=4, column=0, sticky="w", pady=8)
        fin_entry = DateEntry(frame, width=47, date_pattern='yyyy-mm-dd')
        fin_entry.grid(row=4, column=1, pady=8, padx=10)

        tk.Label(frame, text="Selecciona destinos incluidos", bg=FONDO_PRINCIPAL, fg=TEXTO).grid(row=5, column=0, sticky="nw", pady=8)
        destinos = DAODestino().obtener_todo()
        destino_options = [f"{d.get_id()} - {d.get_nombre()}" for d in destinos]
        destinos_listbox = tk.Listbox(frame, selectmode="multiple", width=50, height=10, exportselection=False)
        destinos_listbox.grid(row=5, column=1, pady=8, padx=10)

        for option in destino_options:
            destinos_listbox.insert("end", option)

        if paquete:
            nombre_entry.insert(0, paquete.get_nombre())
            descripcion_entry.insert(0, paquete.get_descripcion())
            precio_entry.insert(0, paquete.get_precio())
            if paquete.get_fecha_inicio():
                inicio_entry.set_date(paquete.get_fecha_inicio())
            if paquete.get_fecha_fin():
                fin_entry.set_date(paquete.get_fecha_fin())
            paquete_destino_ids = {d.get_id() for d in paquete.get_destinos()}
            for index, destino in enumerate(destinos):
                if destino.get_id() in paquete_destino_ids:
                    destinos_listbox.selection_set(index)

        def guardar():
            nombre = nombre_entry.get().strip()
            descripcion = descripcion_entry.get().strip()
            inicio = inicio_entry.get().strip()
            fin = fin_entry.get().strip()
            precio = precio_entry.get().strip()
            seleccionados = destinos_listbox.curselection()
            if not nombre or not precio:
                messagebox.showwarning("Paquete", "Nombre y precio son obligatorios.")
                return
            if not seleccionados:
                messagebox.showwarning("Paquete", "Selecciona al menos un destino para el paquete.")
                return
            try:
                precio_val = float(precio)
            except ValueError:
                messagebox.showwarning("Paquete", "Precio debe ser un número válido.")
                return
            destinos_seleccionados = [destinos[i] for i in seleccionados]

            if paquete:
                paquete.set_nombre(nombre)
                paquete.set_descripcion(descripcion)
                paquete.set_fecha_inicio(inicio)
                paquete.set_fecha_fin(fin)
                paquete.set_precio(precio_val)
                paquete.set_destinos(destinos_seleccionados)
                if DAOPaquete().actualizar(paquete):
                    messagebox.showinfo("Paquete", "Paquete actualizado correctamente.")
                    self._cargar_paquetes()
                    ventana.destroy()
                else:
                    messagebox.showerror("Paquete", "No se pudo actualizar el paquete.")
            else:
                nuevo_paquete = PaqueteTuristico(None, nombre, descripcion, inicio, fin, precio_val, destinos_seleccionados)
                if DAOPaquete().registrar(nuevo_paquete):
                    messagebox.showinfo("Paquete", "Paquete creado correctamente.")
                    self._cargar_paquetes()
                    ventana.destroy()
                else:
                    messagebox.showerror("Paquete", "No se pudo crear el paquete.")

        tk.Button(ventana, text="Guardar", bg=ACCENT_DARK, fg="white", command=guardar, width=14, height=2).pack(pady=18)

    def _vista_reservas(self):
        self._limpiar_contenido()
        tk.Label(self.contenido, text="Gestión de Reservas", bg=FONDO_PRINCIPAL, fg=TEXTO, font=("Arial", 28, "bold")).pack(anchor="w", padx=50, pady=30)

        btn_frame = tk.Frame(self.contenido, bg=FONDO_PRINCIPAL)
        btn_frame.pack(fill="x", padx=50, pady=10)
        tk.Button(btn_frame, text="➕ Nueva Reserva", bg=ACCENT_DARK, fg="white", command=self._nueva_reserva, width=15).pack(side="left", padx=5)
        tk.Button(btn_frame, text="✏️ Editar Reserva", bg=ACCENT, fg="white", command=self._editar_reserva, width=15).pack(side="left", padx=5)
        tk.Button(btn_frame, text="🗑️ Eliminar Reserva", bg=ROJO, fg="white", command=self._eliminar_reserva, width=15).pack(side="left", padx=5)

        columns = ("ID", "Cliente", "Paquete", "Fecha", "Estado")
        tree_frame = tk.Frame(self.contenido, bg=FONDO_PRINCIPAL)
        tree_frame.pack(fill="both", expand=True, padx=50, pady=10)

        self.reserva_tree = ttk.Treeview(tree_frame, columns=columns, show="headings", selectmode="browse")
        for col in columns:
            self.reserva_tree.heading(col, text=col)
            self.reserva_tree.column(col, width=200, anchor="center")
        self.reserva_tree.pack(side="left", fill="both", expand=True)

        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.reserva_tree.yview)
        self.reserva_tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")

        self._cargar_reservas()

    def _cargar_reservas(self):
        self.reserva_tree.delete(*self.reserva_tree.get_children())
        for row in DAOReserva().obtener_todo():
            self.reserva_tree.insert("", "end", values=(row.get("idReserva"), row.get("cliente_nombre"), row.get("paquete_nombre"), row.get("fecha_reserva"), row.get("estado")))

    def _nueva_reserva(self):
        self._abrir_form_reserva()

    def _editar_reserva(self):
        seleccionado = self.reserva_tree.selection()
        if not seleccionado:
            messagebox.showwarning("Editar Reserva", "Selecciona una reserva para editar.")
            return
        item = self.reserva_tree.item(seleccionado)
        valores = item["values"]
        reserva = Reserva(valores[0], None, None, valores[3], valores[4])
        self._abrir_form_reserva(reserva, valores)

    def _eliminar_reserva(self):
        seleccionado = self.reserva_tree.selection()
        if not seleccionado:
            messagebox.showwarning("Eliminar Reserva", "Selecciona una reserva para eliminar.")
            return
        item = self.reserva_tree.item(seleccionado)
        reserva_id = item["values"][0]
        if messagebox.askyesno("Eliminar Reserva", "¿Estás seguro de eliminar esta reserva?"):
            if DAOReserva().eliminar(reserva_id):
                messagebox.showinfo("Reserva", "Reserva eliminada correctamente.")
                self._cargar_reservas()
            else:
                messagebox.showerror("Reserva", "No se pudo eliminar la reserva.")

    def _abrir_form_reserva(self, reserva=None, valores_existentes=None):
        ventana = tk.Toplevel(self.root)
        ventana.title("Reserva")
        ventana.transient(self.root)
        ventana.grab_set()
        ventana.geometry("600x500")
        ventana.configure(bg=FONDO_PRINCIPAL)

        tk.Label(ventana, text="Reserva", bg=FONDO_PRINCIPAL, fg=TEXTO, font=("Arial", 20, "bold")).pack(pady=12)

        frame = tk.Frame(ventana, bg=FONDO_PRINCIPAL)
        frame.pack(fill="both", expand=True, padx=20, pady=10)

        tk.Label(frame, text="Cliente", bg=FONDO_PRINCIPAL, fg=TEXTO).grid(row=0, column=0, sticky="w", pady=8)
        clientes = DAOCliente().obtener_todo()
        cliente_options = [f"{c.get_id()} - {c.get_nombre()}" for c in clientes]
        cliente_combobox = ttk.Combobox(frame, values=cliente_options, state="readonly", width=46)
        cliente_combobox.grid(row=0, column=1, pady=8, padx=10)

        tk.Label(frame, text="Paquete", bg=FONDO_PRINCIPAL, fg=TEXTO).grid(row=1, column=0, sticky="w", pady=8)
        paquetes = DAOPaquete().obtener_todo()
        paquete_options = [f"{p.get_id()} - {p.get_nombre()}" for p in paquetes]
        paquete_combobox = ttk.Combobox(frame, values=paquete_options, state="readonly", width=46)
        paquete_combobox.grid(row=1, column=1, pady=8, padx=10)

        tk.Label(frame, text="Fecha Reserva", bg=FONDO_PRINCIPAL, fg=TEXTO).grid(row=2, column=0, sticky="w", pady=8)
        fecha_entry = DateEntry(frame, width=47, date_pattern='yyyy-mm-dd')
        fecha_entry.grid(row=2, column=1, pady=8, padx=10)

        tk.Label(frame, text="Estado", bg=FONDO_PRINCIPAL, fg=TEXTO).grid(row=3, column=0, sticky="w", pady=8)
        estado_options = ["Pendiente", "Confirmada", "Cancelada", "Pagada"]
        estado_combobox = ttk.Combobox(frame, values=estado_options, state="readonly", width=46)
        estado_combobox.grid(row=3, column=1, pady=8, padx=10)

        if valores_existentes:
            fecha_entry.set_date(valores_existentes[3])
            estado_combobox.set(valores_existentes[4])
            cliente_match = next((opt for opt in cliente_options if opt.split(" - ", 1)[1] == valores_existentes[1]), None)
            paquete_match = next((opt for opt in paquete_options if opt.split(" - ", 1)[1] == valores_existentes[2]), None)
            if cliente_match:
                cliente_combobox.set(cliente_match)
            if paquete_match:
                paquete_combobox.set(paquete_match)
        else:
            estado_combobox.set("Pendiente")
            if clientes:
                cliente_combobox.set(cliente_options[0])
            if paquetes:
                paquete_combobox.set(paquete_options[0])

        def guardar():
            cliente_text = cliente_combobox.get()
            paquete_text = paquete_combobox.get()
            fecha = fecha_entry.get().strip()
            estado = estado_combobox.get().strip()
            if not cliente_text or not paquete_text or not fecha or not estado:
                messagebox.showwarning("Reserva", "Todos los campos son obligatorios.")
                return
            cliente_id = int(cliente_text.split(" - ")[0])
            paquete_id = int(paquete_text.split(" - ")[0])

            if reserva and valores_existentes:
                reserva.set_cliente_id(cliente_id)
                reserva.set_paquete_id(paquete_id)
                reserva.set_fecha_reserva(fecha)
                reserva.set_estado(estado)
                if DAOReserva().actualizar(reserva):
                    messagebox.showinfo("Reserva", "Reserva actualizada correctamente.")
                    self._cargar_reservas()
                    ventana.destroy()
                else:
                    messagebox.showerror("Reserva", "No se pudo actualizar la reserva.")
            else:
                nueva_reserva = Reserva(None, cliente_id, paquete_id, fecha, estado)
                if DAOReserva().registrar(nueva_reserva):
                    messagebox.showinfo("Reserva", "Reserva creada correctamente.")
                    self._cargar_reservas()
                    ventana.destroy()
                else:
                    messagebox.showerror("Reserva", "No se pudo crear la reserva.")

        tk.Button(ventana, text="Guardar", bg=ACCENT_DARK, fg="white", command=guardar, width=14, height=2).pack(pady=18)

    def _vista_exportar(self):
        workbook = Workbook()

        destinos_sheet = workbook.active
        destinos_sheet.title = "Destinos"
        destinos_sheet.append(["ID", "Nombre", "Descripción", "Actividades", "Costo"])
        for destino in DAODestino().obtener_todo():
            destinos_sheet.append([destino.get_id(), destino.get_nombre(), destino.get_descripcion(), destino.get_actividades(), destino.get_costo()])

        paquetes_sheet = workbook.create_sheet("Paquetes")
        paquetes_sheet.append(["ID", "Nombre", "Descripción", "Fecha Inicio", "Fecha Fin", "Precio", "Destinos"])
        for paquete in DAOPaquete().obtener_todo():
            destinos_text = ", ".join([dest.get_nombre() for dest in paquete.get_destinos()])
            paquetes_sheet.append([paquete.get_id(), paquete.get_nombre(), paquete.get_descripcion(), paquete.get_fecha_inicio(), paquete.get_fecha_fin(), paquete.get_precio(), destinos_text])

        reservas_sheet = workbook.create_sheet("Reservas")
        reservas_sheet.append(["ID", "Cliente", "Paquete", "Fecha Reserva", "Estado"])
        for row in DAOReserva().obtener_todo():
            reservas_sheet.append([row.get("idReserva"), row.get("cliente_nombre"), row.get("paquete_nombre"), row.get("fecha_reserva"), row.get("estado")])

        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")], title="Guardar reporte Excel")
        if path:
            workbook.save(path)
            messagebox.showinfo("Exportar", f"Archivo Excel guardado en:\n{path}")
        else:
            messagebox.showwarning("Exportar", "Exportación cancelada.")

    def _panel_cliente(self):
        self._limpiar_sidebar()
        tk.Label(self.sidebar, text="🌍 Viajes Aventura", bg=MENU_BG, fg=TEXTO_CLARO, font=("Arial", 22, "bold")).pack(pady=(30,5))
        nombre = self.usuario_actual.get_nombre() if self.usuario_actual else "Cliente"
        tk.Label(self.sidebar, text=nombre, bg=MENU_BG, fg=ACCENT, font=("Arial", 10, "bold")).pack(pady=(0,20))

        self._boton_menu("Paquetes", self._vista_paquetes_cliente, "🎒")
        self._boton_menu("Mis Reservas", self._vista_mis_reservas, "📋")
        self._boton_menu("Cerrar Sesión", self._cerrar_sesion, "🚪")

        self._vista_paquetes_cliente()

    def _vista_paquetes_cliente(self):
        self._limpiar_contenido()
        tk.Label(self.contenido, text="Paquetes Disponibles", bg=FONDO_PRINCIPAL, fg=TEXTO, font=("Arial", 28, "bold")).pack(anchor="w", padx=50, pady=30)

        columns = ("ID", "Nombre", "Descripción", "Inicio", "Fin", "Precio")
        tree_frame = tk.Frame(self.contenido, bg=FONDO_PRINCIPAL)
        tree_frame.pack(fill="both", expand=True, padx=50, pady=10)

        paquete_tree = ttk.Treeview(tree_frame, columns=columns, show="headings", selectmode="browse")
        for col in columns:
            paquete_tree.heading(col, text=col)
            paquete_tree.column(col, width=180, anchor="center")
        paquete_tree.pack(side="left", fill="both", expand=True)

        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=paquete_tree.yview)
        paquete_tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")

        for paquete in DAOPaquete().obtener_todo():
            paquete_tree.insert("", "end", values=(paquete.get_id(), paquete.get_nombre(), paquete.get_descripcion(), paquete.get_fecha_inicio(), paquete.get_fecha_fin(), paquete.get_precio()))

    def _vista_mis_reservas(self):
        self._limpiar_contenido()
        tk.Label(self.contenido, text="Mis Reservas", bg=FONDO_PRINCIPAL, fg=TEXTO, font=("Arial", 28, "bold")).pack(anchor="w", padx=50, pady=30)

        columns = ("ID", "Paquete", "Fecha", "Estado")
        tree_frame = tk.Frame(self.contenido, bg=FONDO_PRINCIPAL)
        tree_frame.pack(fill="both", expand=True, padx=50, pady=10)

        reserva_tree = ttk.Treeview(tree_frame, columns=columns, show="headings", selectmode="browse")
        for col in columns:
            reserva_tree.heading(col, text=col)
            reserva_tree.column(col, width=200, anchor="center")
        reserva_tree.pack(side="left", fill="both", expand=True)

        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=reserva_tree.yview)
        reserva_tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")

        user_id = self.usuario_actual.get_id() if self.usuario_actual else None
        for row in DAOReserva().obtener_todo():
            if row.get("Cliente_idCliente") == user_id:
                reserva_tree.insert("", "end", values=(row.get("idReserva"), row.get("paquete_nombre"), row.get("fecha_reserva"), row.get("estado")))

    def run(self):
        self.root.mainloop()


if __name__ == "__main__":
    app = App()
    app.run()